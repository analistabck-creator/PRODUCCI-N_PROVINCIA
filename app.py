from __future__ import annotations

import base64
import mimetypes
import os
import shutil
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from flask import (
    Flask,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
except Exception:  # pragma: no cover
    service_account = None
    build = None
    MediaFileUpload = None

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:  # pragma: no cover
    psycopg = None
    dict_row = None


BASE_DIR = Path(__file__).resolve().parent
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
USE_POSTGRES = DATABASE_URL.startswith("postgres://") or DATABASE_URL.startswith("postgresql://")
DATA_DIR = Path(
    os.environ.get("APP_DATA_DIR")
    or os.environ.get("RENDER_DISK_PATH", "")
    or (BASE_DIR / "data")
)
DB_PATH = DATA_DIR / "produccion.db"
UPLOADS_DIR = DATA_DIR / "uploads"
EXPORTS_DIR = DATA_DIR / "exports"
AUTO_EXPORT_PATH = EXPORTS_DIR / "historial_maestro.xlsx"
PDF_EXPORTS_DIR = EXPORTS_DIR / "pdfs"
GOOGLE_SERVICE_ACCOUNT_FILE = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE", "")
GOOGLE_DRIVE_FOLDER_ID = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")
GOOGLE_DRIVE_FILE_NAME = os.environ.get("GOOGLE_DRIVE_FILE_NAME", "historial_produccion_maestro.xlsx")
GOOGLE_DRIVE_SYNC_MODE = os.environ.get("GOOGLE_DRIVE_SYNC_MODE", "full").lower()  # off, excel, full
DEFAULT_PORT = int(os.environ.get("PORT", "10000"))

if not USE_POSTGRES:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
PDF_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "cambia-esta-clave-en-produccion")
app.config["MAX_CONTENT_LENGTH"] = 768 * 1024 * 1024
app.config["PREFERRED_URL_SCHEME"] = "https"

ROLE_OPTIONS = ["administrador", "supervisor", "tecnico", "almacenero"]
STATUS_OPTIONS = ["Borrador", "Enviado", "Revisado", "Aprobado", "Observado", "Cerrado"]
STATUS_BADGES = {
    "Borrador": "neutral",
    "Enviado": "info",
    "Revisado": "warning",
    "Aprobado": "success",
    "Observado": "danger",
    "Cerrado": "dark",
}

FIELDS = [
    {"name": "almacen", "label": "Almacén / sede", "required": True, "type": "text"},
    {"name": "fecha_produccion", "label": "Fecha de producción", "required": True, "type": "date"},
    {"name": "cuenta", "label": "Cuenta", "required": True, "type": "text"},
    {"name": "tipo_maquina", "label": "Tipo de máquina", "required": True, "type": "text"},
    {"name": "marca", "label": "Marca", "required": False, "type": "text"},
    {"name": "modelo", "label": "Modelo", "required": False, "type": "text"},
    {"name": "serie", "label": "Serie", "required": False, "type": "text"},
    {"name": "sticker", "label": "Sticker", "required": False, "type": "text"},
    {"name": "cod_interno_g2m", "label": "Cod. interno G2M", "required": False, "type": "text"},
    {"name": "logo", "label": "Logo", "required": False, "type": "text"},
    {"name": "numero_reporte", "label": "N° de reporte", "required": False, "type": "text"},
    {"name": "estado", "label": "Estado del equipo", "required": True, "type": "select", "options": ["Pendiente", "En proceso", "Completado", "Observado"]},
    {"name": "observacion", "label": "Observación", "required": False, "type": "textarea", "full": True},
    {"name": "lavado", "label": "Lavado", "required": False, "type": "number", "step": "0.01", "min": "0"},
    {"name": "correctivo_menor", "label": "Correctivo menor", "required": False, "type": "number", "step": "0.01", "min": "0"},
    {"name": "correctivo_mayor", "label": "Correctivo mayor", "required": False, "type": "number", "step": "0.01", "min": "0"},
    {"name": "pegado_vinil", "label": "Pegado de vinil", "required": False, "type": "number", "step": "0.01", "min": "0"},
    {"name": "pintura_accesorios", "label": "Pintura accesorios", "required": False, "type": "number", "step": "0.01", "min": "0"},
    {"name": "pintura_cabina", "label": "Pintura de cabina", "required": False, "type": "number", "step": "0.01", "min": "0"},
    {"name": "tec_lavado", "label": "Tec. lavado", "required": False, "type": "text"},
    {"name": "tec_correctivo_menor", "label": "Tec. correctivo menor", "required": False, "type": "text"},
    {"name": "tec_correctivo_mayor", "label": "Tec. correctivo mayor", "required": False, "type": "text"},
    {"name": "tec_pintura", "label": "Tec. pintura", "required": False, "type": "text"},
]

MEDIA_FIELDS = [
    {"name": "foto_placa", "label": "Foto de la placa", "accept": "image/*", "capture": True, "multiple": False},
    {"name": "foto_serie", "label": "Foto de la serie", "accept": "image/*", "capture": True, "multiple": False},
    {"name": "foto_frontal_trasera", "label": "Foto frontal y trasera del equipo", "accept": "image/*", "capture": True, "multiple": True},
    {"name": "foto_laterales", "label": "Foto de los laterales del equipo", "accept": "image/*", "capture": True, "multiple": True},
    {"name": "video_360", "label": "Video 360 interno y externo", "accept": "video/*", "capture": True, "multiple": False},
    {"name": "foto_checklist_firmado", "label": "Foto del checklist firmado", "accept": "image/*", "capture": True, "multiple": False},
    {"name": "foto_reporte_firmado", "label": "Foto del reporte firmado", "accept": "image/*", "capture": True, "multiple": False},
]

FIELD_MAP = {field["name"]: field for field in FIELDS}
MEDIA_FIELD_MAP = {field["name"]: field for field in MEDIA_FIELDS}
NUMERIC_FIELDS = {"lavado", "correctivo_menor", "correctivo_mayor", "pegado_vinil", "pintura_accesorios", "pintura_cabina"}
EXCEL_HEADERS = [
    "N°",
    "ALMACEN",
    "FECHA DE PRODUCCION",
    "CUENTA",
    "TIPO MAQUINA",
    "MARCA",
    "MODELO",
    "SERIE",
    "STIKER",
    "COD. INTERNO G2M",
    "LOGO",
    "N° DE REPORTE",
    "ESTADO",
    "OBSERVACIÓN",
    "LAVADO",
    "CORRECTIVO MENOR",
    "CORRECTIVO MAYOR",
    "PEGADO DE VINIL",
    "PINTURA ACCESORIOS",
    "PINTURA DE CABINA",
    "TEC. LAVADO",
    "TEC. CORRECTIVO MENOR",
    "TEC. CORRECTIVO MAYOR",
    "TEC. PINTURA",
    "FLUJO",
    "SEDE USUARIO",
    "CREADO POR",
    "ROL CREADOR",
    "ULTIMA EDICION POR",
    "TIENE EVIDENCIAS",
    "TOTAL ARCHIVOS",
    "FIRMA TECNICO",
    "FIRMA ALMACENERO",
    "ID DRIVE REGISTRO",
    "ULTIMO SYNC DRIVE",
    "ESTADO SYNC DRIVE",
    "CREADO",
    "ACTUALIZADO",
]
ALLOWED_FILE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif", ".mp4", ".mov", ".avi", ".mkv", ".3gp", ".m4v", ".pdf"}


# --------------------------- Database helpers ---------------------------
class DBCursorWrapper:
    def __init__(self, cursor, db_type: str, inserted_id: int | None = None):
        self.cursor = cursor
        self.db_type = db_type
        self.lastrowid = inserted_id if inserted_id is not None else getattr(cursor, "lastrowid", None)

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()

    def __iter__(self):
        return iter(self.cursor)


class DBWrapper:
    def __init__(self, conn, db_type: str):
        self.conn = conn
        self.db_type = db_type

    def execute(self, query: str, params: tuple | list | None = None):
        params = tuple(params or ())
        if self.db_type == "postgres":
            q = query.replace("?", "%s")
            cur = self.conn.cursor()
            inserted_id = None
            normalized = q.strip().rstrip(";")
            upper = normalized.upper()
            if upper.startswith("INSERT INTO") and " RETURNING " not in upper:
                q = normalized + " RETURNING id"
                cur.execute(q, params)
                returned = cur.fetchone()
                inserted_id = returned["id"] if returned else None
            else:
                cur.execute(q, params)
            return DBCursorWrapper(cur, self.db_type, inserted_id)
        cur = self.conn.execute(query, params)
        return DBCursorWrapper(cur, self.db_type)

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def close(self):
        self.conn.close()


def get_db() -> DBWrapper:
    if "db" not in g:
        if USE_POSTGRES:
            if psycopg is None:
                raise RuntimeError("psycopg no está instalado, pero DATABASE_URL está configurado.")
            conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
            g.db = DBWrapper(conn, "postgres")
        else:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            g.db = DBWrapper(conn, "sqlite")
    return g.db


@app.teardown_appcontext
def close_db(exc) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def iso_now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def human_dt(value: str | None) -> str:
    if not value:
        return "-"
    try:
        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y %H:%M")
    except Exception:
        return value


def get_table_columns(db, table: str) -> list[str]:
    if getattr(db, "db_type", "sqlite") == "postgres":
        rows = db.execute(
            "SELECT column_name FROM information_schema.columns WHERE table_schema = 'public' AND table_name = ?",
            (table,),
        ).fetchall()
        return [row["column_name"] for row in rows]
    rows = db.execute(f"PRAGMA table_info({table})").fetchall()
    return [row[1] for row in rows]


def init_db() -> None:
    if USE_POSTGRES:
        if psycopg is None:
            raise RuntimeError("Falta instalar psycopg para PostgreSQL.")
        raw = psycopg.connect(DATABASE_URL, row_factory=dict_row)
        db = DBWrapper(raw, "postgres")
        bool_default = "BOOLEAN NOT NULL DEFAULT TRUE"
        id_col = "BIGSERIAL PRIMARY KEY"
        real_col = "DOUBLE PRECISION"
    else:
        raw = sqlite3.connect(DB_PATH)
        raw.row_factory = sqlite3.Row
        raw.execute("PRAGMA foreign_keys = ON")
        db = DBWrapper(raw, "sqlite")
        bool_default = "INTEGER NOT NULL DEFAULT 1"
        id_col = "INTEGER PRIMARY KEY AUTOINCREMENT"
        real_col = "REAL"

    db.execute(f'''
        CREATE TABLE IF NOT EXISTS users (
            id {id_col},
            nombre TEXT NOT NULL,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            sede TEXT NOT NULL,
            is_active {bool_default},
            created_at TEXT NOT NULL
        )
    ''')
    db.execute(f'''
        CREATE TABLE IF NOT EXISTS production_records (
            id {id_col},
            almacen TEXT NOT NULL,
            fecha_produccion TEXT NOT NULL,
            cuenta TEXT NOT NULL,
            tipo_maquina TEXT NOT NULL,
            marca TEXT,
            modelo TEXT,
            serie TEXT,
            sticker TEXT,
            cod_interno_g2m TEXT,
            logo TEXT,
            numero_reporte TEXT,
            estado TEXT NOT NULL,
            observacion TEXT,
            lavado {real_col},
            correctivo_menor {real_col},
            correctivo_mayor {real_col},
            pegado_vinil {real_col},
            pintura_accesorios {real_col},
            pintura_cabina {real_col},
            tec_lavado TEXT,
            tec_correctivo_menor TEXT,
            tec_correctivo_mayor TEXT,
            tec_pintura TEXT,
            workflow_status TEXT NOT NULL DEFAULT 'Borrador',
            submitted_at TEXT,
            reviewed_at TEXT,
            approved_at TEXT,
            observed_at TEXT,
            closed_at TEXT,
            created_by_user_id BIGINT,
            created_by_name TEXT,
            created_by_role TEXT,
            created_by_sede TEXT,
            last_edited_by_user_id BIGINT,
            last_edited_by_name TEXT,
            firma_tecnico_path TEXT,
            firma_almacenero_path TEXT,
            drive_record_folder_id TEXT,
            drive_last_sync_at TEXT,
            drive_sync_status TEXT,
            drive_last_sync_message TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(created_by_user_id) REFERENCES users(id),
            FOREIGN KEY(last_edited_by_user_id) REFERENCES users(id)
        )
    ''')
    db.execute(f'''
        CREATE TABLE IF NOT EXISTS media_files (
            id {id_col},
            record_id BIGINT NOT NULL,
            media_key TEXT NOT NULL,
            original_name TEXT NOT NULL,
            stored_name TEXT NOT NULL,
            relative_path TEXT NOT NULL,
            mime_type TEXT,
            drive_file_id TEXT,
            drive_web_link TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(record_id) REFERENCES production_records(id) ON DELETE CASCADE
        )
    ''')
    db.execute(f'''
        CREATE TABLE IF NOT EXISTS activity_log (
            id {id_col},
            record_id BIGINT NOT NULL,
            action TEXT NOT NULL,
            detail TEXT,
            actor_name TEXT NOT NULL,
            actor_role TEXT NOT NULL,
            actor_sede TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(record_id) REFERENCES production_records(id) ON DELETE CASCADE
        )
    ''')
    db.execute(f'''
        CREATE TABLE IF NOT EXISTS drive_sync_log (
            id {id_col},
            record_id BIGINT,
            sync_scope TEXT NOT NULL,
            status TEXT NOT NULL,
            message TEXT,
            drive_file_id TEXT,
            drive_link TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(record_id) REFERENCES production_records(id) ON DELETE CASCADE
        )
    ''')

    def ensure_column(table: str, name: str, ddl: str) -> None:
        cols = get_table_columns(db, table)
        if name not in cols:
            db.execute(f"ALTER TABLE {table} ADD COLUMN {ddl}")

    ensure_column("production_records", "drive_record_folder_id", "drive_record_folder_id TEXT")
    ensure_column("production_records", "drive_last_sync_at", "drive_last_sync_at TEXT")
    ensure_column("production_records", "drive_sync_status", "drive_sync_status TEXT")
    ensure_column("production_records", "drive_last_sync_message", "drive_last_sync_message TEXT")
    ensure_column("media_files", "drive_file_id", "drive_file_id TEXT")
    ensure_column("media_files", "drive_web_link", "drive_web_link TEXT")

    demo_users = [
        ("Administrador", "admin", "admin123", "administrador", "Central"),
        ("Supervisor", "supervisor", "super123", "supervisor", "Central"),
        ("Tecnico Demo", "tecnico", "tec123", "tecnico", "Provincia 1"),
        ("Almacenero Demo", "almacen", "alm123", "almacenero", "Provincia 1"),
    ]
    for nombre, username, password, role, sede in demo_users:
        exists = db.execute("SELECT 1 FROM users WHERE username = ?", (username,)).fetchone()
        if not exists:
            db.execute(
                "INSERT INTO users (nombre, username, password_hash, role, sede, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (nombre, username, generate_password_hash(password), role, sede, iso_now()),
            )
    db.commit()
    db.close()


# --------------------------- Auth & template helpers ---------------------------
def current_user() -> dict[str, Any] | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id = ? AND is_active = 1", (user_id,)).fetchone()
    return dict(row) if row else None


@app.before_request
def load_user() -> None:
    g.user = current_user()


@app.context_processor
def inject_globals() -> dict[str, Any]:
    return {
        "current_user": g.get("user"),
        "workflow_options": STATUS_OPTIONS,
        "role_options": ROLE_OPTIONS,
        "sync_modes": ["off", "excel", "full"],
    }


def login_required() -> None:
    if not g.user:
        abort(401)


def require_role(*roles: str) -> None:
    login_required()
    if g.user["role"] not in roles:
        abort(403)


# --------------------------- Parsing & file helpers ---------------------------
def parse_form_data(form) -> tuple[dict[str, Any], list[str]]:
    payload: dict[str, Any] = {}
    errors: list[str] = []
    for field in FIELDS:
        raw = (form.get(field["name"]) or "").strip()
        if field["required"] and not raw:
            errors.append(f"El campo {field['label']} es obligatorio.")
        if field["type"] == "number":
            if raw == "":
                payload[field["name"]] = None
            else:
                try:
                    payload[field["name"]] = float(raw)
                except ValueError:
                    errors.append(f"El campo {field['label']} debe ser numérico.")
                    payload[field["name"]] = None
        else:
            payload[field["name"]] = raw or None
    workflow_status = (form.get("workflow_status") or "Borrador").strip()
    if workflow_status not in STATUS_OPTIONS:
        workflow_status = "Borrador"
    payload["workflow_status"] = workflow_status
    return payload, errors


def safe_ext(filename: str) -> str:
    ext = Path(filename or "").suffix.lower()
    return ext if ext in ALLOWED_FILE_EXTENSIONS else ".bin"


def save_uploaded_file(file_storage, subfolder: str) -> str | None:
    if not file_storage or not file_storage.filename:
        return None
    filename = secure_filename(file_storage.filename)
    ext = safe_ext(filename)
    target_dir = UPLOADS_DIR / subfolder
    target_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{Path(filename).stem[:40]}_{uuid4().hex[:10]}{ext}"
    path = target_dir / stored_name
    file_storage.save(path)
    return str(path.relative_to(BASE_DIR)).replace("\\", "/")


def save_signature(data_url: str, subfolder: str, prefix: str) -> str | None:
    if not data_url or "," not in data_url:
        return None
    header, encoded = data_url.split(",", 1)
    if "base64" not in header:
        return None
    try:
        binary = base64.b64decode(encoded)
    except Exception:
        return None
    target_dir = UPLOADS_DIR / subfolder
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{prefix}_{uuid4().hex[:12]}.png"
    path = target_dir / filename
    path.write_bytes(binary)
    return str(path.relative_to(BASE_DIR)).replace("\\", "/")


def attach_media(db: Any, record_id: int, files_by_key: dict[str, list[Any]]) -> None:
    now = iso_now()
    for media_key, files in files_by_key.items():
        for file_storage in files:
            rel_path = save_uploaded_file(file_storage, f"record_{record_id}/{media_key}")
            if not rel_path:
                continue
            mime_type = file_storage.mimetype or mimetypes.guess_type(file_storage.filename)[0] or "application/octet-stream"
            db.execute(
                """
                INSERT INTO media_files (record_id, media_key, original_name, stored_name, relative_path, mime_type, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    record_id,
                    media_key,
                    file_storage.filename,
                    Path(rel_path).name,
                    rel_path,
                    mime_type,
                    now,
                ),
            )


def record_media_map(db: Any, record_id: int) -> dict[str, list[Any]]:
    rows = db.execute(
        "SELECT * FROM media_files WHERE record_id = ? ORDER BY id DESC",
        (record_id,),
    ).fetchall()
    grouped: dict[str, list[Any]] = defaultdict(list)
    for row in rows:
        grouped[row["media_key"]].append(row)
    return grouped


def media_count(db: Any, record_id: int) -> int:
    return db.execute("SELECT COUNT(*) FROM media_files WHERE record_id = ?", (record_id,)).fetchone()[0]


def log_activity(db: Any, record_id: int, action: str, detail: str = "") -> None:
    user = current_user() or {"nombre": "Sistema", "role": "sistema", "sede": "Central"}
    db.execute(
        "INSERT INTO activity_log (record_id, action, detail, actor_name, actor_role, actor_sede, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (record_id, action, detail, user["nombre"], user["role"], user["sede"], iso_now()),
    )


def drive_enabled() -> bool:
    return bool(
        GOOGLE_DRIVE_SYNC_MODE != "off"
        and GOOGLE_SERVICE_ACCOUNT_FILE
        and GOOGLE_DRIVE_FOLDER_ID
        and service_account
        and build
        and MediaFileUpload
    )


def get_drive_service():
    if not drive_enabled():
        return None
    creds = service_account.Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def drive_safe_name(name: str) -> str:
    return (name or "sin_nombre").replace("'", "").replace("/", "-").replace("\\", "-").strip()[:120]


def log_drive_sync(db: Any, record_id: int | None, sync_scope: str, status: str, message: str = "", drive_file_id: str | None = None, drive_link: str | None = None) -> None:
    db.execute(
        "INSERT INTO drive_sync_log (record_id, sync_scope, status, message, drive_file_id, drive_link, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (record_id, sync_scope, status, message, drive_file_id, drive_link, iso_now()),
    )


def ensure_drive_folder(service, name: str, parent_id: str) -> tuple[str, str | None]:
    q = f"name = '{drive_safe_name(name)}' and '{parent_id}' in parents and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
    result = service.files().list(q=q, fields="files(id, webViewLink)", pageSize=1).execute()
    files = result.get("files", [])
    if files:
        return files[0]["id"], files[0].get("webViewLink")
    created = service.files().create(
        body={"name": drive_safe_name(name), "parents": [parent_id], "mimeType": "application/vnd.google-apps.folder"},
        fields="id, webViewLink",
    ).execute()
    return created["id"], created.get("webViewLink")


def upload_file_to_drive(service, filepath: Path, parent_id: str, file_name: str | None = None, mimetype: str | None = None) -> tuple[str, str | None]:
    file_name = drive_safe_name(file_name or filepath.name)
    q = f"name = '{file_name}' and '{parent_id}' in parents and trashed = false"
    existing = service.files().list(q=q, fields="files(id, webViewLink)", pageSize=1).execute().get("files", [])
    media = MediaFileUpload(str(filepath), mimetype=mimetype or mimetypes.guess_type(filepath.name)[0] or "application/octet-stream", resumable=False)
    if existing:
        updated = service.files().update(fileId=existing[0]["id"], media_body=media, fields="id, webViewLink").execute()
        return updated["id"], updated.get("webViewLink")
    created = service.files().create(body={"name": file_name, "parents": [parent_id]}, media_body=media, fields="id, webViewLink").execute()
    return created["id"], created.get("webViewLink")


def build_pdf(record: Any, media_map: dict[str, list[Any]]) -> BytesIO:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Reporte de Producción", styles["Title"]))
    story.append(Spacer(1, 0.25 * cm))
    info_rows = [["Campo", "Valor"]]
    for field in FIELDS:
        info_rows.append([field["label"], str(record[field["name"]] or "-")])
    info_rows.extend([
        ["Flujo", record["workflow_status"] or "-"],
        ["Creado por", f"{record['created_by_name'] or '-'} ({record['created_by_role'] or '-'})"],
        ["Sede usuario", record["created_by_sede"] or "-"],
        ["Última edición", record["last_edited_by_name"] or "-"],
        ["Creado", record["created_at"] or "-"],
        ["Actualizado", record["updated_at"] or "-"],
    ])
    tbl = Table(info_rows, colWidths=[5.2 * cm, 12.4 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B66")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D7E2")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FC")]),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("Resumen de evidencias", styles["Heading2"]))
    evidence_rows = [["Tipo", "Cantidad"]]
    for media_field in MEDIA_FIELDS:
        evidence_rows.append([media_field["label"], str(len(media_map.get(media_field["name"], [])))])
    etbl = Table(evidence_rows, colWidths=[11 * cm, 4 * cm])
    etbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D5E8C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D0D7E2")),
    ]))
    story.append(etbl)
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("Firmas registradas", styles["Heading2"]))
    for label, path_key in [("Firma técnico", "firma_tecnico_path"), ("Firma almacenero", "firma_almacenero_path")]:
        story.append(Paragraph(label, styles["Heading4"]))
        sign_path = record[path_key]
        if sign_path and (BASE_DIR / sign_path).exists():
            story.append(Image(str(BASE_DIR / sign_path), width=6 * cm, height=2.4 * cm))
        else:
            story.append(Paragraph("No registrada", styles["Normal"]))
        story.append(Spacer(1, 0.2 * cm))
    doc.build(story)
    buffer.seek(0)
    return buffer


def write_record_pdf_to_disk(record: Any, db: Any) -> Path:
    pdf_path = PDF_EXPORTS_DIR / f"reporte_produccion_{record['id']}.pdf"
    pdf_bytes = build_pdf(record, record_media_map(db, record["id"]))
    pdf_path.write_bytes(pdf_bytes.read())
    return pdf_path


def sync_drive_excel(service, db: Any) -> tuple[str | None, str | None]:
    if not AUTO_EXPORT_PATH.exists():
        return None, None
    file_id, link = upload_file_to_drive(
        service,
        AUTO_EXPORT_PATH,
        GOOGLE_DRIVE_FOLDER_ID,
        file_name=GOOGLE_DRIVE_FILE_NAME,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    log_drive_sync(db, None, "excel_maestro", "success", "Excel maestro sincronizado", file_id, link)
    return file_id, link


def sync_record_to_drive(db: Any, record_id: int) -> tuple[bool, str]:
    if not drive_enabled():
        return False, "Google Drive no está configurado."
    record = db.execute("SELECT * FROM production_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        return False, "Registro no encontrado."
    service = get_drive_service()
    if service is None:
        return False, "No se pudo conectar con Drive."
    try:
        root_folder_id = GOOGLE_DRIVE_FOLDER_ID
        records_folder_id, _ = ensure_drive_folder(service, "registros_produccion", root_folder_id)
        record_folder_name = f"registro_{record['id']}_{record['almacen'] or 'sede'}_{record['serie'] or record['numero_reporte'] or 'sinreferencia'}"
        record_folder_id, record_folder_link = ensure_drive_folder(service, record_folder_name, records_folder_id)

        # Upload PDF snapshot
        pdf_path = write_record_pdf_to_disk(record, db)
        upload_file_to_drive(service, pdf_path, record_folder_id, file_name=pdf_path.name, mimetype="application/pdf")

        # Upload signatures
        for label, path_key in [("firma_tecnico", "firma_tecnico_path"), ("firma_almacenero", "firma_almacenero_path")]:
            sign_rel = record[path_key]
            if sign_rel and (BASE_DIR / sign_rel).exists():
                sign_path = BASE_DIR / sign_rel
                upload_file_to_drive(service, sign_path, record_folder_id, file_name=f"{label}_{record['id']}.png", mimetype="image/png")

        # Upload media evidence
        media_rows = db.execute("SELECT * FROM media_files WHERE record_id = ? ORDER BY id", (record_id,)).fetchall()
        for item in media_rows:
            file_path = BASE_DIR / item["relative_path"]
            if not file_path.exists():
                continue
            evid_folder_id, _ = ensure_drive_folder(service, item["media_key"], record_folder_id)
            drive_file_id, drive_link = upload_file_to_drive(service, file_path, evid_folder_id, file_name=item["stored_name"], mimetype=item["mime_type"])
            db.execute(
                "UPDATE media_files SET drive_file_id = ?, drive_web_link = ? WHERE id = ?",
                (drive_file_id, drive_link, item["id"]),
            )

        db.execute(
            "UPDATE production_records SET drive_record_folder_id = ?, drive_last_sync_at = ?, drive_sync_status = ?, drive_last_sync_message = ? WHERE id = ?",
            (record_folder_id, iso_now(), "success", "Sincronizado correctamente con Google Drive", record_id),
        )
        log_drive_sync(db, record_id, "registro", "success", "Registro sincronizado con Drive", record_folder_id, record_folder_link)
        return True, "Registro sincronizado con Google Drive."
    except Exception as exc:
        db.execute(
            "UPDATE production_records SET drive_last_sync_at = ?, drive_sync_status = ?, drive_last_sync_message = ? WHERE id = ?",
            (iso_now(), "error", str(exc), record_id),
        )
        log_drive_sync(db, record_id, "registro", "error", str(exc))
        return False, str(exc)


def sync_drive_if_configured(db: Any, record_id: int | None = None) -> None:
    if not drive_enabled():
        return
    service = get_drive_service()
    if service is None:
        return
    try:
        sync_drive_excel(service, db)
        if GOOGLE_DRIVE_SYNC_MODE == "full" and record_id is not None:
            sync_record_to_drive(db, record_id)
        db.commit()
    except Exception as exc:
        log_drive_sync(db, record_id, "general", "error", str(exc))
        db.commit()


def export_master_excel(db: Any) -> Path:
    rows = db.execute("SELECT * FROM production_records ORDER BY id DESC").fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"
    ws.append(EXCEL_HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="B7C9E2")
    for idx, title in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_index, record in enumerate(rows, start=2):
        total_files = media_count(db, record["id"])
        ws.append([
            record["id"], record["almacen"], record["fecha_produccion"], record["cuenta"], record["tipo_maquina"],
            record["marca"], record["modelo"], record["serie"], record["sticker"], record["cod_interno_g2m"],
            record["logo"], record["numero_reporte"], record["estado"], record["observacion"], record["lavado"],
            record["correctivo_menor"], record["correctivo_mayor"], record["pegado_vinil"], record["pintura_accesorios"],
            record["pintura_cabina"], record["tec_lavado"], record["tec_correctivo_menor"], record["tec_correctivo_mayor"],
            record["tec_pintura"], record["workflow_status"], record["created_by_sede"], record["created_by_name"],
            record["created_by_role"], record["last_edited_by_name"], "Sí" if total_files else "No", total_files,
            "Sí" if record["firma_tecnico_path"] else "No", "Sí" if record["firma_almacenero_path"] else "No",
            record["drive_record_folder_id"], record["drive_last_sync_at"], record["drive_sync_status"],
            record["created_at"], record["updated_at"],
        ])
        for col in range(1, len(EXCEL_HEADERS) + 1):
            ws.cell(row=row_index, column=col).border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.cell(row=row_index, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    widths = [8,18,16,14,18,14,14,18,12,18,12,16,12,24,12,14,14,14,18,18,16,18,18,14,14,16,18,14,18,14,12,12,12,18,18,14,20,20]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    wb.save(AUTO_EXPORT_PATH)
    return AUTO_EXPORT_PATH


def decorate_record(record: Any, db: Any) -> dict[str, Any]:
    data = dict(record)
    data["media_count"] = media_count(db, record["id"])
    data["created_at_human"] = human_dt(record["created_at"])
    data["updated_at_human"] = human_dt(record["updated_at"])
    data["drive_last_sync_human"] = human_dt(record["drive_last_sync_at"])
    data["status_badge"] = STATUS_BADGES.get(record["workflow_status"], "neutral")
    data["sync_badge"] = "success" if record["drive_sync_status"] == "success" else ("danger" if record["drive_sync_status"] == "error" else "neutral")
    return data


def can_edit_record(user: dict[str, Any], record: Any) -> bool:
    return user["role"] in {"administrador", "supervisor"} or record["created_by_user_id"] == user["id"]


def can_change_workflow(user: dict[str, Any]) -> bool:
    return user["role"] in {"administrador", "supervisor", "almacenero"}


# --------------------------- Routes ---------------------------
@app.route("/")
def home():
    if not g.user:
        return redirect(url_for("login"))
    return redirect(url_for("dashboard"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            flash(f"Bienvenido, {user['nombre']}.", "success")
            return redirect(url_for("dashboard"))
        flash("Usuario o contraseña inválidos.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada.", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    login_required()
    db = get_db()
    base_query = "SELECT * FROM production_records"
    params: list[Any] = []
    if g.user["role"] in {"tecnico", "almacenero"}:
        base_query += " WHERE created_by_sede = ?"
        params.append(g.user["sede"])
    rows = db.execute(base_query + " ORDER BY id DESC", params).fetchall()
    decorated = [decorate_record(row, db) for row in rows]
    status_counts = Counter(row["workflow_status"] for row in rows)
    sede_counts = Counter(row["almacen"] or "Sin sede" for row in rows)
    top_tecnicos = Counter((row["tec_lavado"] or row["tec_pintura"] or row["created_by_name"] or "Sin técnico") for row in rows)
    recent = decorated[:8]
    sync_summary = Counter(row["drive_sync_status"] or "pendiente" for row in rows)
    drive_log = db.execute("SELECT * FROM drive_sync_log ORDER BY id DESC LIMIT 10").fetchall()
    metrics = {
        "total": len(rows),
        "aprobados": status_counts.get("Aprobado", 0),
        "observados": status_counts.get("Observado", 0),
        "cerrados": status_counts.get("Cerrado", 0),
        "con_evidencias": sum(1 for r in decorated if r["media_count"] > 0),
        "drive_ok": sync_summary.get("success", 0),
        "drive_error": sync_summary.get("error", 0),
    }
    return render_template("dashboard.html", metrics=metrics, status_counts=status_counts, sede_counts=sede_counts, top_tecnicos=top_tecnicos, recent=recent, drive_log=drive_log)


@app.route("/records")
def records_history():
    login_required()
    db = get_db()
    where = []
    params: list[Any] = []
    q = (request.args.get("q") or "").strip()
    workflow = (request.args.get("workflow") or "").strip()
    almacen = (request.args.get("almacen") or "").strip()
    if g.user["role"] in {"tecnico", "almacenero"}:
        where.append("created_by_sede = ?")
        params.append(g.user["sede"])
    if q:
        where.append("(cuenta LIKE ? OR serie LIKE ? OR numero_reporte LIKE ? OR modelo LIKE ?)")
        wildcard = f"%{q}%"
        params.extend([wildcard, wildcard, wildcard, wildcard])
    if workflow:
        where.append("workflow_status = ?")
        params.append(workflow)
    if almacen:
        where.append("almacen LIKE ?")
        params.append(f"%{almacen}%")
    sql = "SELECT * FROM production_records"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()
    records = [decorate_record(row, db) for row in rows]
    return render_template("history.html", records=records, filters={"q": q, "workflow": workflow, "almacen": almacen})


@app.route("/records/new", methods=["GET", "POST"])
def record_create():
    login_required()
    db = get_db()
    if request.method == "POST":
        payload, errors = parse_form_data(request.form)
        user = g.user
        if payload["workflow_status"] in {"Revisado", "Aprobado", "Cerrado"} and not can_change_workflow(user):
            payload["workflow_status"] = "Borrador"
        if not request.form.get("signature_tecnico"):
            errors.append("La firma del técnico es obligatoria.")
        if not request.form.get("signature_almacenero"):
            errors.append("La firma del almacenero es obligatoria.")
        if errors:
            for error in errors:
                flash(error, "danger")
            return render_template("record_form.html", fields=FIELDS, media_fields=MEDIA_FIELDS, values=request.form, mode="create")
        now = iso_now()
        firma_tecnico_path = save_signature(request.form.get("signature_tecnico", ""), "signatures", f"tecnico_{user['id']}")
        firma_almacenero_path = save_signature(request.form.get("signature_almacenero", ""), "signatures", f"almacenero_{user['id']}")
        cursor = db.execute(
            f"""
            INSERT INTO production_records (
                {', '.join(field['name'] for field in FIELDS)},
                workflow_status, submitted_at, reviewed_at, approved_at, observed_at, closed_at,
                created_by_user_id, created_by_name, created_by_role, created_by_sede,
                last_edited_by_user_id, last_edited_by_name,
                firma_tecnico_path, firma_almacenero_path,
                created_at, updated_at
            ) VALUES (
                {', '.join(['?'] * len(FIELDS))},
                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
            )
            """,
            [payload[field["name"]] for field in FIELDS]
            + [
                payload["workflow_status"],
                now if payload["workflow_status"] == "Enviado" else None,
                now if payload["workflow_status"] == "Revisado" else None,
                now if payload["workflow_status"] == "Aprobado" else None,
                now if payload["workflow_status"] == "Observado" else None,
                now if payload["workflow_status"] == "Cerrado" else None,
                user["id"], user["nombre"], user["role"], user["sede"], user["id"], user["nombre"],
                firma_tecnico_path, firma_almacenero_path, now, now,
            ],
        )
        record_id = cursor.lastrowid
        files_payload = {field["name"]: request.files.getlist(field["name"]) for field in MEDIA_FIELDS}
        attach_media(db, record_id, files_payload)
        log_activity(db, record_id, "Creación", f"Registro creado con flujo {payload['workflow_status']}")
        db.commit()
        export_master_excel(db)
        sync_drive_if_configured(db, record_id)
        flash("Registro creado correctamente.", "success")
        return redirect(url_for("record_detail", record_id=record_id))
    return render_template("record_form.html", fields=FIELDS, media_fields=MEDIA_FIELDS, values={}, mode="create")


@app.route("/records/<int:record_id>")
def record_detail(record_id: int):
    login_required()
    db = get_db()
    record = db.execute("SELECT * FROM production_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        abort(404)
    if g.user["role"] in {"tecnico", "almacenero"} and record["created_by_sede"] != g.user["sede"]:
        abort(403)
    media_map = record_media_map(db, record_id)
    activity = db.execute("SELECT * FROM activity_log WHERE record_id = ? ORDER BY id DESC", (record_id,)).fetchall()
    drive_log = db.execute("SELECT * FROM drive_sync_log WHERE record_id = ? ORDER BY id DESC LIMIT 12", (record_id,)).fetchall()
    return render_template("detail.html", record=decorate_record(record, db), fields=FIELDS, media_fields=MEDIA_FIELDS, media_map=media_map, activity=activity, drive_log=drive_log)


@app.route("/records/<int:record_id>/edit", methods=["GET", "POST"])
def record_edit(record_id: int):
    login_required()
    db = get_db()
    record = db.execute("SELECT * FROM production_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        abort(404)
    if not can_edit_record(g.user, record):
        abort(403)
    if request.method == "POST":
        payload, errors = parse_form_data(request.form)
        if errors:
            for error in errors:
                flash(error, "danger")
            return render_template("record_form.html", fields=FIELDS, media_fields=MEDIA_FIELDS, values=request.form, mode="edit", record=record)
        now = iso_now()
        extra_updates = []
        extra_values = []
        if request.form.get("signature_tecnico"):
            extra_updates.append("firma_tecnico_path = ?")
            extra_values.append(save_signature(request.form.get("signature_tecnico", ""), "signatures", f"tecnico_edit_{g.user['id']}"))
        if request.form.get("signature_almacenero"):
            extra_updates.append("firma_almacenero_path = ?")
            extra_values.append(save_signature(request.form.get("signature_almacenero", ""), "signatures", f"almacenero_edit_{g.user['id']}"))
        db.execute(
            f"UPDATE production_records SET {', '.join(f'{field['name']} = ?' for field in FIELDS)}, workflow_status = ?, last_edited_by_user_id = ?, last_edited_by_name = ?, updated_at = ?{', ' + ', '.join(extra_updates) if extra_updates else ''} WHERE id = ?",
            [payload[field["name"]] for field in FIELDS] + [payload["workflow_status"], g.user["id"], g.user["nombre"], now] + extra_values + [record_id],
        )
        files_payload = {field["name"]: request.files.getlist(field["name"]) for field in MEDIA_FIELDS}
        attach_media(db, record_id, files_payload)
        log_activity(db, record_id, "Edición", "Registro actualizado")
        db.commit()
        export_master_excel(db)
        sync_drive_if_configured(db, record_id)
        flash("Registro actualizado correctamente.", "success")
        return redirect(url_for("record_detail", record_id=record_id))
    values = dict(record)
    return render_template("record_form.html", fields=FIELDS, media_fields=MEDIA_FIELDS, values=values, mode="edit", record=record)


@app.route("/records/<int:record_id>/workflow", methods=["POST"])
def record_workflow(record_id: int):
    login_required()
    user = g.user
    if not can_change_workflow(user):
        abort(403)
    new_status = (request.form.get("workflow_status") or "").strip()
    if new_status not in STATUS_OPTIONS:
        flash("Estado de flujo inválido.", "danger")
        return redirect(url_for("record_detail", record_id=record_id))
    db = get_db()
    record = db.execute("SELECT * FROM production_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        abort(404)
    now = iso_now()
    stamp_map = {
        "Enviado": ("submitted_at", now),
        "Revisado": ("reviewed_at", now),
        "Aprobado": ("approved_at", now),
        "Observado": ("observed_at", now),
        "Cerrado": ("closed_at", now),
    }
    clauses = ["workflow_status = ?", "last_edited_by_user_id = ?", "last_edited_by_name = ?", "updated_at = ?"]
    params: list[Any] = [new_status, user["id"], user["nombre"], now]
    if new_status in stamp_map:
        clauses.append(f"{stamp_map[new_status][0]} = ?")
        params.append(stamp_map[new_status][1])
    params.append(record_id)
    db.execute(f"UPDATE production_records SET {', '.join(clauses)} WHERE id = ?", params)
    log_activity(db, record_id, "Cambio de flujo", f"Flujo cambiado a {new_status}")
    db.commit()
    export_master_excel(db)
    sync_drive_if_configured(db, record_id)
    flash("Flujo actualizado.", "success")
    return redirect(url_for("record_detail", record_id=record_id))


@app.route("/records/<int:record_id>/delete-media/<int:media_id>", methods=["POST"])
def delete_media(record_id: int, media_id: int):
    login_required()
    db = get_db()
    record = db.execute("SELECT * FROM production_records WHERE id = ?", (record_id,)).fetchone()
    if not record or not can_edit_record(g.user, record):
        abort(403)
    media = db.execute("SELECT * FROM media_files WHERE id = ? AND record_id = ?", (media_id, record_id)).fetchone()
    if not media:
        abort(404)
    file_path = BASE_DIR / media["relative_path"]
    if file_path.exists():
        file_path.unlink()
    db.execute("DELETE FROM media_files WHERE id = ?", (media_id,))
    db.execute("UPDATE production_records SET updated_at = ?, last_edited_by_user_id = ?, last_edited_by_name = ? WHERE id = ?", (iso_now(), g.user["id"], g.user["nombre"], record_id))
    log_activity(db, record_id, "Eliminación de evidencia", media["original_name"])
    db.commit()
    export_master_excel(db)
    sync_drive_if_configured(db, record_id)
    flash("Archivo eliminado.", "info")
    return redirect(url_for("record_detail", record_id=record_id))


@app.route("/records/<int:record_id>/pdf")
def record_pdf(record_id: int):
    login_required()
    db = get_db()
    record = db.execute("SELECT * FROM production_records WHERE id = ?", (record_id,)).fetchone()
    if not record:
        abort(404)
    if g.user["role"] in {"tecnico", "almacenero"} and record["created_by_sede"] != g.user["sede"]:
        abort(403)
    pdf = build_pdf(record, record_media_map(db, record_id))
    return send_file(pdf, as_attachment=True, download_name=f"reporte_produccion_{record_id}.pdf", mimetype="application/pdf")


@app.route("/records/<int:record_id>/sync-drive", methods=["POST"])
def record_sync_drive(record_id: int):
    login_required()
    require_role("administrador", "supervisor", "almacenero")
    db = get_db()
    export_master_excel(db)
    ok, message = sync_record_to_drive(db, record_id)
    sync_drive_if_configured(db, None)
    db.commit()
    flash(message, "success" if ok else "danger")
    return redirect(url_for("record_detail", record_id=record_id))


@app.route("/drive/sync-all", methods=["POST"])
def drive_sync_all():
    login_required()
    require_role("administrador", "supervisor")
    db = get_db()
    export_master_excel(db)
    service = get_drive_service()
    if service is None:
        flash("Drive no está configurado todavía.", "danger")
        return redirect(url_for("drive_center"))
    sync_drive_excel(service, db)
    rows = db.execute("SELECT id FROM production_records ORDER BY id DESC LIMIT 50").fetchall()
    synced = 0
    errors = 0
    for row in rows:
        ok, _ = sync_record_to_drive(db, row["id"])
        synced += int(ok)
        errors += int(not ok)
    db.commit()
    flash(f"Sincronización masiva terminada. Correctos: {synced}. Con error: {errors}.", "info")
    return redirect(url_for("drive_center"))


@app.route("/records/export")
def records_export():
    login_required()
    db = get_db()
    path = export_master_excel(db)
    sync_drive_if_configured(db, None)
    return send_file(path, as_attachment=True, download_name=path.name)


@app.route("/uploads/<path:filename>")
def uploaded_file(filename: str):
    login_required()
    return send_from_directory(BASE_DIR, filename)


@app.route("/admin/users", methods=["GET", "POST"])
def admin_users():
    login_required()
    require_role("administrador")
    db = get_db()
    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        role = (request.form.get("role") or "").strip()
        sede = (request.form.get("sede") or "").strip()
        if not all([nombre, username, password, role, sede]) or role not in ROLE_OPTIONS:
            flash("Completa todos los datos del usuario.", "danger")
        else:
            try:
                db.execute(
                    "INSERT INTO users (nombre, username, password_hash, role, sede, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (nombre, username, generate_password_hash(password), role, sede, iso_now()),
                )
                db.commit()
                flash("Usuario creado.", "success")
            except Exception as exc:
                if "unique" not in str(exc).lower() and "duplicate" not in str(exc).lower():
                    raise
                flash("El nombre de usuario ya existe.", "danger")
    users = db.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
    return render_template("users.html", users=users)


@app.route("/drive")
def drive_center():
    login_required()
    require_role("administrador", "supervisor")
    db = get_db()
    latest_logs = db.execute("SELECT * FROM drive_sync_log ORDER BY id DESC LIMIT 40").fetchall()
    records = db.execute("SELECT * FROM production_records ORDER BY id DESC LIMIT 20").fetchall()
    config_status = {
        "enabled": drive_enabled(),
        "service_account": GOOGLE_SERVICE_ACCOUNT_FILE or "No configurado",
        "folder_id": GOOGLE_DRIVE_FOLDER_ID or "No configurado",
        "excel_name": GOOGLE_DRIVE_FILE_NAME,
        "mode": GOOGLE_DRIVE_SYNC_MODE,
    }
    return render_template("drive.html", latest_logs=latest_logs, records=[decorate_record(r, db) for r in records], config_status=config_status)




@app.route("/healthz")
def healthz():
    try:
        db = get_db()
        db.execute("SELECT 1").fetchone()
        return {"ok": True, "database": "postgres" if USE_POSTGRES else "sqlite"}, 200
    except Exception as exc:
        return {"ok": False, "error": str(exc)}, 500


# --------------------------- Error handlers ---------------------------
@app.errorhandler(401)
def unauthorized(error):
    return redirect(url_for("login"))


@app.errorhandler(403)
def forbidden(error):
    return render_template("error.html", title="Acceso denegado", message="No tienes permiso para entrar a esta sección."), 403


@app.errorhandler(404)
def not_found(error):
    return render_template("error.html", title="No encontrado", message="No se encontró el recurso solicitado."), 404


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=DEFAULT_PORT, debug=os.environ.get("FLASK_DEBUG", "0") == "1")
else:
    init_db()
